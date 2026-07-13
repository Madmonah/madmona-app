# -*- coding: utf-8 -*-
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

S = r'E:\madmona-app\scripts'
leads = []
for f in ['dz-leads.json', 'dz-leads2.json']:
    p = os.path.join(S, f)
    if os.path.exists(p):
        with open(p, encoding='utf-8') as fh:
            try: leads += json.load(fh)
            except Exception: pass

# تصنيف: مالك فرد ولا مكتب صغير
OFFICE = ['future villa','abrag','smart assets','onyx','estate','properties','realty','group']
def kind(n):
    ln = (n or '').lower()
    return 'مكتب/وسيط صغير' if any(k in ln for k in OFFICE) else 'مالك فرد ✅'

# دمج بالرقم
seen = {}
for l in leads:
    ph = l.get('phone','')
    if not ph: continue
    if ph in seen:
        seen[ph]['units'] += 1
        seen[ph]['max_m'] = max(seen[ph]['max_m'], l.get('price_m',0))
        continue
    seen[ph] = {
        'name': l.get('name',''),
        'phone': ph,
        'type': kind(l.get('name','')),
        'max_m': l.get('price_m',0),
        'units': 1,
        'area': l.get('area') or l.get('kind') or 'الساحل',
        'title': (l.get('title') or '')[:80],
        'url': l.get('url',''),
    }

rows = sorted(seen.values(), key=lambda r: -r['max_m'])

wb = Workbook(); ws = wb.active; ws.title = 'Leads A+'
ws.sheet_view.rightToLeft = True

HDR = ['#','الاسم','رقم الموبايل','النوع','أعلى سعر (مليون ج)','عدد الوحدات','المنطقة','الوحدة','رابط الإعلان','الحالة','ملاحظات']
ws.append(HDR)
hf = PatternFill('solid', fgColor='0F5132'); hfont = Font(bold=True, color='FFFFFF', size=11)
thin = Side(style='thin', color='CCCCCC'); bd = Border(left=thin,right=thin,top=thin,bottom=thin)
for c in ws[1]:
    c.fill = hf; c.font = hfont; c.alignment = Alignment(horizontal='center', vertical='center'); c.border = bd
ws.freeze_panes = 'A2'

gold = PatternFill('solid', fgColor='FFF3CD')
green = PatternFill('solid', fgColor='D1E7DD')

for i, r in enumerate(rows, 1):
    ws.append([i, r['name'], r['phone'], r['type'], r['max_m'], r['units'], r['area'], r['title'], r['url'], '', ''])
    row = ws.max_row
    for c in ws[row]: c.border = bd; c.alignment = Alignment(vertical='center', wrap_text=False)
    if r['type'].startswith('مالك'):
        for c in ws[row]: c.fill = green
    if r['max_m'] >= 50:
        ws.cell(row=row, column=5).fill = gold
        ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=9).hyperlink = r['url']
    ws.cell(row=row, column=9).value = 'فتح الإعلان'
    ws.cell(row=row, column=9).font = Font(color='0563C1', underline='single')

for col, w in zip('ABCDEFGHIJK', [5,26,16,18,20,12,18,55,14,14,26]):
    ws.column_dimensions[col].width = w
ws.auto_filter.ref = f'A1:K{ws.max_row}'

# تبويب المصدر والمنهجية
ws2 = wb.create_sheet('المصدر')
ws2.sheet_view.rightToLeft = True
info = [
    ['المصدر','Dubizzle مصر (OLX) — إعلانات منشورة علناً'],
    ['الفلتر','فيلات وشاليهات الساحل + فيلات القاهرة والجيزة'],
    ['حد السعر','٢٠ مليون جنيه فأكتر (الساحل) · ٣٠ مليون فأكتر (القاهرة/الجيزة)'],
    ['الاستبعاد','اتشال كل حساب عليه علامة Verified Business وكل اسم فيه كلمة شركة/عقارات/بروكر'],
    ['الأرقام','مأخوذة من زرار «Show Phone Number» اللي المعلن نفسه حاطه عشان الناس تكلمه'],
    ['ملحوظة مهمة','اللي مكتوب قصاده «مالك فرد» = شخص بيبيع وحدته بنفسه. اللي «مكتب/وسيط صغير» = بيبيع أكتر من وحدة، ينفع شراكة بروكر مش عميل.'],
    ['',''],
    ['إزاي تستخدمه','ابدأ بالأخضر (ملاك أفراد) — دول A+ فعلاً وعندهم كاش. اعرض عليهم مضمونة كوسيط مضمون للبيع أو كفرصة شراء في المشاريع الجديدة.'],
]
for r in info: ws2.append(r)
for c in ws2['A']: c.font = Font(bold=True)
ws2.column_dimensions['A'].width = 18; ws2.column_dimensions['B'].width = 95
for row in ws2.iter_rows():
    for c in row: c.alignment = Alignment(wrap_text=True, vertical='center')

out = r'E:\madmona-app\scripts\madmona-leads-A-plus.xlsx'
wb.save(out)
print('SAVED', out, '| rows =', len(rows))
for r in rows: print(r['type'], '|', r['name'], '|', r['phone'], '|', r['max_m'], 'M')
