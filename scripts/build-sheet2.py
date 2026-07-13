# -*- coding: utf-8 -*-
import json, os, io, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

S = r'E:\madmona-app\scripts'
def load(f):
    p = os.path.join(S, f)
    if not os.path.exists(p): return []
    try:
        with open(p, encoding='utf-8') as fh: return json.load(fh)
    except Exception: return []

props = load('leads-v5.json')      # عقارات بيع (فلتر صارم)
rent  = load('leads-rent.json')    # إيجار فيلات
cars  = load('leads-cars.json')    # عربيات لوكشري

wb = Workbook()
thin = Side(style='thin', color='D0D0D0'); bd = Border(left=thin,right=thin,top=thin,bottom=thin)
HDRF = PatternFill('solid', fgColor='0F5132'); HF = Font(bold=True, color='FFFFFF', size=11)
GOLD = PatternFill('solid', fgColor='FFF3CD')

def sheet(ws, rows, hdr, widths, money_col=None, gold_at=None):
    ws.sheet_view.rightToLeft = True
    ws.append(hdr)
    for c in ws[1]:
        c.fill = HDRF; c.font = HF; c.border = bd
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'
    for i, r in enumerate(rows, 1):
        ws.append([i] + r['cells'] + ['', ''])
        row = ws.max_row
        for c in ws[row]: c.border = bd; c.alignment = Alignment(vertical='center')
        ws.cell(row=row, column=3).font = Font(bold=True)   # الرقم
        if gold_at and r['val'] >= gold_at:
            ws.cell(row=row, column=money_col).fill = GOLD
            ws.cell(row=row, column=money_col).font = Font(bold=True)
        lc = len(hdr) - 2
        ws.cell(row=row, column=lc).hyperlink = r['url']
        ws.cell(row=row, column=lc).value = 'فتح الإعلان'
        ws.cell(row=row, column=lc).font = Font(color='0563C1', underline='single')
    for col, w in zip('ABCDEFGHIJ', widths):
        ws.column_dimensions[col].width = w
    ws.auto_filter.ref = f'A1:{chr(64+len(hdr))}{ws.max_row}'

# --- 1) عقارات للبيع ---
ws1 = wb.active; ws1.title = 'عقارات A+ (بيع)'
rows = []
seen = set()
for l in sorted(props, key=lambda x: -x.get('price_m', 0)):
    if l['phone'] in seen: continue
    seen.add(l['phone'])
    rows.append({'cells': [l['name'], l['phone'], l['price_m'], l.get('kind',''),
                           l.get('activeAds',''), (l.get('title') or '')[:70]],
                 'url': l['url'], 'val': l.get('price_m', 0)})
sheet(ws1, rows, ['#','الاسم','الموبايل','السعر (مليون ج)','النوع','عدد إعلاناته','الوحدة','الإعلان','الحالة','ملاحظات'],
      [5,24,15,17,16,13,50,13,13,24], money_col=4, gold_at=50)

# --- 2) عربيات لوكشري (بس لو فيه داتا) ---
if cars:
    ws2 = wb.create_sheet('عربيات لوكشري')
    rows = []; seen = set()
    for l in sorted(cars, key=lambda x: -x.get('price_m', 0)):
        if l['phone'] in seen: continue
        seen.add(l['phone'])
        rows.append({'cells': [l['name'], l['phone'], l['price_m'], l.get('activeAds',''), (l.get('title') or '')[:70]],
                     'url': l['url'], 'val': l.get('price_m', 0)})
    sheet(ws2, rows, ['#','الاسم','الموبايل','السعر (مليون ج)','عدد إعلاناته','العربية','الإعلان','الحالة','ملاحظات'],
          [5,24,15,17,13,50,13,13,24], money_col=4, gold_at=8)

# --- 3) إيجار فيلات ---
ws3 = wb.create_sheet('إيجار فيلات')
rows = []; seen = set()
for l in sorted(rent, key=lambda x: -x.get('price_m', 0)):
    if l['phone'] in seen: continue
    seen.add(l['phone'])
    egp = int(l.get('price_m', 0) * 1e6)
    rows.append({'cells': [l['name'], l['phone'], egp, l.get('kind',''), l.get('activeAds',''), (l.get('title') or '')[:70]],
                 'url': l['url'], 'val': egp})
sheet(ws3, rows, ['#','الاسم','الموبايل','الإيجار (ج)','النوع','عدد إعلاناته','الوحدة','الإعلان','الحالة','ملاحظات'],
      [5,24,15,15,17,13,50,13,13,24], money_col=4, gold_at=50000)

# --- 4) المنهجية ---
ws4 = wb.create_sheet('المنهجية')
ws4.sheet_view.rightToLeft = True
info = [
    ['المصدر','Dubizzle مصر (OLX) — إعلانات منشورة علناً، والأرقام من زرار «Show Phone Number» اللي المعلن نفسه حاطه عشان الناس تكلمه'],
    ['',''],
    ['⭐ الفلتر (٣ طبقات)','ده اللي بيفرّق بين المالك والشركة:'],
    ['١) عدد الإعلانات','أي معلن عنده أكتر من ٣ إعلانات = شركة → اتشال. (المالك عنده وحدة أو اتنين)'],
    ['٢) الاسم','أي اسم فيه كلمة عقارات/شركة/بروكر/Real Estate/Properties وخلافه → اتشال'],
    ['٣) Verified Business','أي حساب عليه علامة تجارية من Dubizzle → اتشال'],
    ['',''],
    ['حد السعر','عقارات: ٢٠ مليون+ (الساحل) · ٣٠ مليون+ (القاهرة والجيزة)'],
    ['',''],
    ['⚠️ حقيقة مهمة','Dubizzle الساحل حوالي ٩٠٪ منه بروكرز. الملاك الأفراد قليلين فعلاً — فالعدد صغير لأن الفلتر صارم عن قصد، مش لأن فيه مشكلة. اتفحص ٣١٧ إعلان طلع منهم ٩ ملاك حقيقيين.'],
    ['',''],
    ['⚠️ تبويب الإيجارات','دي إيجارات صيفية عادية (٦–٣٠ ألف جنيه) — ملاك فيلات فعلاً، بس مش بالضرورة شريحة A+ زي تبويب البيع. استخدمهم كملاك وحدات مش كعملاء أثرياء.'],
    ['',''],
    ['إزاي تستخدمه','دول ناس حاطة أرقامها بنفسها عشان حد يكلمها. اعرض عليهم مضمونة كوسيط مضمون للبيع، أو كفرصة شراء في المشاريع الجديدة اللي في البورصة.'],
    ['المصفّر بالذهبي','أعلى شريحة — ابدأ بيهم.'],
]
for r in info: ws4.append(r)
for c in ws4['A']: c.font = Font(bold=True)
ws4.column_dimensions['A'].width = 20; ws4.column_dimensions['B'].width = 100
for row in ws4.iter_rows():
    for c in row: c.alignment = Alignment(wrap_text=True, vertical='center')

out = os.path.join(S, 'madmona-leads-A-plus.xlsx')
wb.save(out)
print('SAVED', out)
print('props=', ws1.max_row-1, 'cars=', ws2.max_row-1, 'rent=', ws3.max_row-1)
