# -*- coding: utf-8 -*-
import json, os, io, sys, csv
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

S = r'E:\madmona-app\scripts'
OUTDIR = r'C:\Users\solutions\AppData\Roaming\Claude\local-agent-mode-sessions\b4c8f83c-bfd2-4d1f-a7fa-bfb88a9ebea8\7327d46d-6790-4c13-8b1a-9aee6e2447a3\local_4f4ba56b-2cfb-4d83-8fe8-624b6fdf9a43\outputs'

def load(f):
    p = os.path.join(S, f)
    if not os.path.exists(p): return []
    try:
        with open(p, encoding='utf-8') as fh: return json.load(fh)
    except Exception: return []

props = load('leads-v5.json')
rent  = load('leads-rent.json')
cars  = load('leads-cars.json')

wb = Workbook()
thin = Side(style='thin', color='D0D0D0'); bd = Border(left=thin,right=thin,top=thin,bottom=thin)
HDRF = PatternFill('solid', fgColor='0F5132'); HF = Font(bold=True, color='FFFFFF', size=11)
GOLD = PatternFill('solid', fgColor='FFF3CD')

def build(ws, hdr, rows, widths, gold_at):
    ws.sheet_view.rightToLeft = True
    ws.append(hdr)
    for c in ws[1]:
        c.fill = HDRF; c.font = HF; c.border = bd
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'
    for i, r in enumerate(rows, 1):
        ws.append([i] + r['cells'])
        row = ws.max_row
        for c in ws[row]:
            c.border = bd; c.alignment = Alignment(vertical='center')
        ws.cell(row=row, column=3).font = Font(bold=True)
        if gold_at and r['val'] >= gold_at:
            ws.cell(row=row, column=4).fill = GOLD
            ws.cell(row=row, column=4).font = Font(bold=True)
    for col, w in zip('ABCDEFGHI', widths):
        ws.column_dimensions[col].width = w

# 1) عقارات
ws = wb.active; ws.title = 'عقارات للبيع'
rows, seen = [], set()
for l in sorted(props, key=lambda x: -x.get('price_m', 0)):
    if l['phone'] in seen: continue
    seen.add(l['phone'])
    rows.append({'cells': [l['name'], l['phone'], l['price_m'], l.get('kind',''),
                           l.get('activeAds',''), (l.get('title') or '')[:70], l['url']],
                 'val': l.get('price_m', 0)})
build(ws, ['م','الاسم','الموبايل','السعر بالمليون','النوع','عدد اعلاناته','الوحدة','رابط الاعلان'],
      rows, [5,24,15,16,16,13,52,60], 50)
props_rows = rows

# 2) عربيات
if cars:
    ws = wb.create_sheet('عربيات لوكشري')
    rows2, seen = [], set()
    for l in sorted(cars, key=lambda x: -x.get('price_m', 0)):
        if l['phone'] in seen: continue
        seen.add(l['phone'])
        rows2.append({'cells': [l['name'], l['phone'], l['price_m'], l.get('activeAds',''),
                                (l.get('title') or '')[:70], l['url']], 'val': l.get('price_m', 0)})
    build(ws, ['م','الاسم','الموبايل','السعر بالمليون','عدد اعلاناته','العربية','رابط الاعلان'],
          rows2, [5,24,15,16,13,52,60], 8)

# 3) ايجارات
ws = wb.create_sheet('ايجار فيلات')
rows3, seen = [], set()
for l in sorted(rent, key=lambda x: -x.get('price_m', 0)):
    if l['phone'] in seen: continue
    seen.add(l['phone'])
    egp = int(l.get('price_m', 0) * 1e6)
    rows3.append({'cells': [l['name'], l['phone'], egp, l.get('kind',''), l.get('activeAds',''),
                            (l.get('title') or '')[:70], l['url']], 'val': egp})
build(ws, ['م','الاسم','الموبايل','الايجار بالجنيه','النوع','عدد اعلاناته','الوحدة','رابط الاعلان'],
      rows3, [5,24,15,16,16,13,52,60], 50000)

# 4) منهجية
ws = wb.create_sheet('المنهجية')
ws.sheet_view.rightToLeft = True
for r in [
 ['المصدر','Dubizzle مصر - اعلانات منشورة علنا. الارقام من زرار Show Phone Number اللي المعلن حاطه بنفسه'],
 ['',''],
 ['الفلتر - 3 طبقات','ده اللي بيفرق بين المالك والشركة'],
 ['1) عدد الاعلانات','اي معلن عنده اكتر من 3 اعلانات = شركة - اتشال'],
 ['2) الاسم','اي اسم فيه كلمة عقارات او شركة او بروكر او Real Estate - اتشال'],
 ['3) Verified Business','اي حساب عليه علامة تجارية من Dubizzle - اتشال'],
 ['',''],
 ['حد السعر','الساحل 20 مليون فاكتر - القاهرة والجيزة 30 مليون فاكتر'],
 ['',''],
 ['حقيقة مهمة','Dubizzle الساحل حوالي 90% منه بروكرز. اتفحص 317 اعلان طلع منهم 9 ملاك حقيقيين بس. العدد صغير لان الفلتر صارم عن قصد مش لان فيه مشكلة'],
 ['',''],
 ['تنبيه على الايجارات','دي ايجارات صيفية عادية 6 لـ 30 الف جنيه. ملاك فيلات فعلا بس مش شريحة A+ زي تبويب البيع'],
 ['',''],
 ['ازاي تستخدمه','دول ناس حاطة ارقامها بنفسها عشان حد يكلمها. اعرض عليهم مضمونة كوسيط مضمون للبيع او كفرصة شراء في مشاريع البورصة'],
 ['المصفر بالذهبي','اعلى شريحة - ابدا بيهم'],
]:
    ws.append(r)
for c in ws['A']: c.font = Font(bold=True)
ws.column_dimensions['A'].width = 22; ws.column_dimensions['B'].width = 100
for row in ws.iter_rows():
    for c in row: c.alignment = Alignment(wrap_text=True, vertical='center')

out = os.path.join(OUTDIR, 'madmona-leads.xlsx')
wb.save(out)
print('XLSX', out)

# CSV احتياطي
cpath = os.path.join(OUTDIR, 'madmona-leads.csv')
with open(cpath, 'w', encoding='utf-8-sig', newline='') as fh:
    w = csv.writer(fh)
    w.writerow(['القسم','الاسم','الموبايل','السعر','النوع','عدد اعلاناته','الوحدة','الرابط'])
    for r in props_rows: w.writerow(['عقار للبيع'] + r['cells'])
    for r in rows3:      w.writerow(['ايجار فيلا'] + r['cells'])
print('CSV', cpath)
print('props', len(props_rows), 'rent', len(rows3), 'cars', len(cars))
