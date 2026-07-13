# -*- coding: utf-8 -*-
import json, os, io, sys, csv
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

S = r'E:\madmona-app\scripts'
OUTDIR = r'C:\Users\solutions\AppData\Roaming\Claude\local-agent-mode-sessions\b4c8f83c-bfd2-4d1f-a7fa-bfb88a9ebea8\7327d46d-6790-4c13-8b1a-9aee6e2447a3\local_4f4ba56b-2cfb-4d83-8fe8-624b6fdf9a43\outputs'

def load(f):
    p = os.path.join(S, f)
    if not os.path.exists(p): return []
    try:
        with open(p, encoding='utf-8') as fh: return json.load(fh)
    except Exception: return []

props = load('leads-v5.json')
rent  = load('leads-rent.json')
cars  = load('leads-cars.json')

# vibe csv
vibe = []
vp = os.path.join(S, 'vibe.csv')
if os.path.exists(vp):
    with open(vp, encoding='utf-8', errors='replace') as fh:
        for r in csv.DictReader(fh):
            vibe.append({
                'name': r.get('prospect_full_name',''),
                'title': (r.get('prospect_job_title','') or '')[:60],
                'company': r.get('prospect_company_name',''),
                'city': r.get('prospect_city','') or r.get('prospect_region_name','') or '',
                'li': r.get('prospect_linkedin',''),
                'site': r.get('prospect_company_website',''),
            })

wb = Workbook()
thin = Side(style='thin', color='D0D0D0'); bd = Border(left=thin,right=thin,top=thin,bottom=thin)
HDRF = PatternFill('solid', fgColor='0F5132'); HF = Font(bold=True, color='FFFFFF', size=11)
BLUE = PatternFill('solid', fgColor='0B5394')
GOLD = PatternFill('solid', fgColor='FFF3CD')

def build(ws, hdr, rows, widths, gold_at=None, gold_col=4, fill=HDRF):
    ws.sheet_view.rightToLeft = True
    ws.append(hdr)
    for c in ws[1]:
        c.fill = fill; c.font = HF; c.border = bd
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'
    for i, r in enumerate(rows, 1):
        ws.append([i] + r['cells'])
        row = ws.max_row
        for c in ws[row]: c.border = bd; c.alignment = Alignment(vertical='center')
        if gold_at and r.get('val', 0) >= gold_at:
            ws.cell(row=row, column=gold_col).fill = GOLD
            ws.cell(row=row, column=gold_col).font = Font(bold=True)
    for col, w in zip('ABCDEFGHI', widths):
        ws.column_dimensions[col].width = w

# 1) عقارات
ws = wb.active; ws.title = 'عقارات للبيع'
rows, seen = [], set()
for l in sorted(props, key=lambda x: -x.get('price_m', 0)):
    if l['phone'] in seen: continue
    seen.add(l['phone'])
    rows.append({'cells': [l['name'], l['phone'], l['price_m'], l.get('kind',''),
                           l.get('activeAds',''), (l.get('title') or '')[:70], l['url']],
                 'val': l.get('price_m', 0)})
build(ws, ['م','الاسم','الموبايل','السعر بالمليون','النوع','عدد اعلاناته','الوحدة','رابط الاعلان'],
      rows, [5,24,15,16,16,13,52,60], 50)
props_rows = rows

# 2) عربيات
if cars:
    ws = wb.create_sheet('عربيات لوكشري')
    r2, seen = [], set()
    for l in sorted(cars, key=lambda x: -x.get('price_m', 0)):
        if l['phone'] in seen: continue
        seen.add(l['phone'])
        r2.append({'cells': [l['name'], l['phone'], l['price_m'], l.get('activeAds',''),
                             (l.get('title') or '')[:70], l['url']], 'val': l.get('price_m', 0)})
    build(ws, ['م','الاسم','الموبايل','السعر بالمليون','عدد اعلاناته','العربية','رابط الاعلان'],
          r2, [5,24,15,16,13,52,60], 8)

# 3) ايجارات
ws = wb.create_sheet('ايجار فيلات')
r3, seen = [], set()
for l in sorted(rent, key=lambda x: -x.get('price_m', 0)):
    if l['phone'] in seen: continue
    seen.add(l['phone'])
    egp = int(l.get('price_m', 0) * 1e6)
    r3.append({'cells': [l['name'], l['phone'], egp, l.get('kind',''), l.get('activeAds',''),
                         (l.get('title') or '')[:70], l['url']], 'val': egp})
build(ws, ['م','الاسم','الموبايل','الايجار بالجنيه','النوع','عدد اعلاناته','الوحدة','رابط الاعلان'],
      r3, [5,24,15,16,16,13,52,60], 50000)

# 4) كبار التنفيذيين
if vibe:
    ws = wb.create_sheet('كبار التنفيذيين')
    r4 = [{'cells': [v['name'], v['title'], v['company'], v['city'], v['li'], v['site']]} for v in vibe]
    build(ws, ['م','الاسم','المنصب','الشركة','المدينة','لينكدإن','موقع الشركة'],
          r4, [5,26,50,26,14,58,26], fill=BLUE)

# 5) منهجية
ws = wb.create_sheet('المنهجية')
ws.sheet_view.rightToLeft = True
for r in [
 ['تبويب عقارات وايجارات وعربيات',''],
 ['المصدر','Dubizzle مصر - اعلانات منشورة علنا. الارقام من زرار Show Phone Number اللي المعلن حاطه بنفسه عشان حد يكلمه'],
 ['الفلتر - 3 طبقات','1) اي معلن عنده اكتر من 3 اعلانات = شركة واتشال. 2) اي اسم فيه كلمة عقارات او شركة او بروكر او Real Estate اتشال. 3) اي حساب عليه علامة Verified Business اتشال'],
 ['حد السعر','الساحل 20 مليون فاكتر - القاهرة والجيزة 30 مليون فاكتر'],
 ['حقيقة مهمة','Dubizzle الساحل حوالي 90% منه بروكرز. اتفحص 317 اعلان طلع منهم 9 ملاك حقيقيين بس'],
 ['تنبيه على الايجارات','دي ايجارات صيفية عادية 6 لـ 30 الف جنيه. ملاك فيلات فعلا بس مش شريحة A+ زي تبويب البيع'],
 ['',''],
 ['تبويب كبار التنفيذيين',''],
 ['المصدر','Vibe Prospecting - Explorium. داتا B2B من لينكدإن'],
 ['مين دول','198 من كبار التنفيذيين في مصر - CEOs ومؤسسين وشركاء. شركات زي Breadfast و Beltone و BCG و Mobica'],
 ['تحذير مهم جدا','الليست دي فيها اسم ومنصب ولينكدإن بس - مفيش موبايل ولا ايميل. اضافة بيانات الاتصال بتكلف 1188 كريدت والرصيد مش كفاية'],
 ['فرق جوهري','الفرق بين التبويبين: ناس Dubizzle حاطة ارقامها بنفسها وعايزة حد يكلمها. كبار التنفيذيين دول مطلبوش حد يكلمهم - فمكالمة باردة ليهم نسبة ردها واطية'],
 ['ازاي تستخدمهم صح','مش كليست تليفونات - كليست حسابات مستهدفة. توصلهم بمقدمة من حد يعرفهم او بمحتوى او على لينكدإن. مش بمكالمة باردة'],
]:
    ws.append(r)
for c in ws['A']: c.font = Font(bold=True)
ws.column_dimensions['A'].width = 24; ws.column_dimensions['B'].width = 105
for row in ws.iter_rows():
    for c in row: c.alignment = Alignment(wrap_text=True, vertical='center')

out = os.path.join(OUTDIR, 'madmona-leads.xlsx')
wb.save(out)
print('SAVED', out)
print('props', len(props_rows), 'rent', len(r3), 'cars', len(cars), 'vibe', len(vibe))
