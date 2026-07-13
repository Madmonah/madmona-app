# -*- coding: utf-8 -*-
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
p = r'E:\madmona-app\scripts\madmona-leads-A-plus.xlsx'
print('exists', os.path.exists(p), 'size', os.path.getsize(p))
wb = load_workbook(p)
print('sheets', wb.sheetnames)
for n in wb.sheetnames:
    ws = wb[n]
    print(n, 'rows', ws.max_row, 'cols', ws.max_column)
