# -*- coding: utf-8 -*-
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
p = r'C:\Users\solutions\AppData\Roaming\Claude\local-agent-mode-sessions\b4c8f83c-bfd2-4d1f-a7fa-bfb88a9ebea8\7327d46d-6790-4c13-8b1a-9aee6e2447a3\local_4f4ba56b-2cfb-4d83-8fe8-624b6fdf9a43\outputs\madmona-leads.xlsx'
wb = load_workbook(p)
print('OPENS OK | sheets:', wb.sheetnames)
ws = wb['عقارات للبيع']
for r in ws.iter_rows(min_row=1, max_row=4, values_only=True):
    print(' | '.join(str(x)[:22] if x is not None else '' for x in r))
